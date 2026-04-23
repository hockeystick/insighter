"""Build fixtures/taxonomy_seed.json from the Plus-Hub xlsx.

Reads the "02 Audience Development" sheet and produces one Cluster per L1
and one CapabilityItem per L3 task, mapping the metadata columns to the
model fields. L2 subcluster is preserved inside each item's description
field (no schema change).

Run from project root with the xlsx present on disk:
  .venv/bin/python scripts/build_taxonomy_fixture.py
"""
from __future__ import annotations

import json
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "Plus-Hub-Service Layer Tasklists [INTERNAL, WIP] (1).xlsx"
OUT = REPO / "fixtures" / "taxonomy_seed.json"
SHEET = "02 Audience Development"

PRIORITY_MAP = {"High": "H", "Medium": "M", "Low": "L"}
YN_MAP = {"Yes": True, "No": False}
PHASE_MAP = {"Advisory": "advisory", "Activation": "activation", "Resources": "resources"}
ASSESSMENT_MAP = {
    "Interview": "interview",
    "Survey": "survey",
    "FTP": "file_transfer",
    "Auto": "automated",
}


def _norm(v):
    if v is None:
        return ""
    return str(v).strip()


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: xlsx not found at {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    header = [c.value for c in ws[1]]
    rows = [
        dict(zip(header, r))
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[1] is not None
    ]

    # Preserve first-appearance cluster order from the sheet.
    cluster_order: "OrderedDict[str, int]" = OrderedDict()
    for r in rows:
        name = _norm(r["L1: Cluster"])
        if name and name not in cluster_order:
            cluster_order[name] = len(cluster_order) + 1

    fixtures = []
    for name, order in cluster_order.items():
        fixtures.append({
            "model": "capabilities.cluster",
            "pk": order,
            "fields": {
                "desk": "audience_development",
                "name": name,
                "description": "",
                "order": order,
            },
        })

    # Items: pk = cluster_order * 1000 + running index within cluster.
    per_cluster_counter: dict[int, int] = {}
    for r in rows:
        cluster_name = _norm(r["L1: Cluster"])
        cluster_pk = cluster_order[cluster_name]
        per_cluster_counter[cluster_pk] = per_cluster_counter.get(cluster_pk, 0) + 1
        idx = per_cluster_counter[cluster_pk]

        task = _norm(r["L3: Task"])
        subcluster = _norm(r["L2: Subcluster"])

        priority = PRIORITY_MAP.get(_norm(r["Priority"]), "M")
        baseline = YN_MAP.get(_norm(r["Baseline?"]), False)
        templateable = YN_MAP.get(_norm(r["Template-able?"]), False)
        phase = PHASE_MAP.get(_norm(r["Phase"]), "advisory")
        assessment = ASSESSMENT_MAP.get(_norm(r["Assessment"]), "interview")
        secondary = _norm(r["Secondary"])

        description_parts = []
        if subcluster:
            description_parts.append(f"Subcluster: {subcluster}")
        notes = _norm(r.get("Notes"))
        if notes:
            description_parts.append(f"Notes: {notes}")

        fixtures.append({
            "model": "capabilities.capabilityitem",
            "pk": cluster_pk * 1000 + idx,
            "fields": {
                "cluster": cluster_pk,
                "name": task[:128],
                "description": "\n".join(description_parts),
                "priority": priority,
                "baseline": baseline,
                "template_able": templateable,
                "phase": phase,
                "assessment_method": assessment,
                "cross_desk_dep": secondary,
                "order": idx,
            },
        })

    OUT.write_text(json.dumps(fixtures, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {len(fixtures)} objects to {OUT}")
    print(f"  clusters: {len(cluster_order)}")
    print(f"  items:    {sum(per_cluster_counter.values())}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
